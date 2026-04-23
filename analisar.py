"""
Analisa a integridade financeira do demonstrativos.xlsx gerado por scan_gemini.py.

Uso:
    pixi run python analisar.py                      # analisa demonstrativos.xlsx
    pixi run python analisar.py outro_arquivo.xlsx   # analisa outro arquivo
"""

import argparse
import sys
from pathlib import Path

import openpyxl


def carregar_evolucao(caminho: str) -> tuple[list[str], dict]:
    """Carrega a aba Evolução e retorna (meses, data_dict)."""
    wb = openpyxl.load_workbook(caminho, data_only=True)
    if "Evolução" not in wb.sheetnames:
        print("❌ Aba 'Evolução' não encontrada.")
        sys.exit(1)

    ws = wb["Evolução"]
    rows = [list(r) for r in ws.iter_rows(min_row=1, values_only=True)]
    meses = rows[0][2:]

    data = {"_wb_path": caminho}
    for r in rows[1:]:
        tipo = r[0] or ""
        desc = r[1] or ""
        vals = {m: r[2 + i] for i, m in enumerate(meses)}
        data[desc] = {"tipo": tipo, "vals": vals}

    return list(meses), data


def v(data, desc, mes):
    return data.get(desc, {}).get("vals", {}).get(mes)


# ── Checks ───────────────────────────────────────────────────────

def check_saldo_continuidade(meses, data):
    """Saldo Anterior[m] deve ser igual a Saldo Atual[m-1]."""
    erros = []
    for i in range(1, len(meses)):
        sa = v(data, "Saldo Anterior", meses[i])
        prev = v(data, "Saldo Atual", meses[i - 1])
        if sa is not None and prev is not None and abs(sa - prev) > 0.02:
            erros.append(
                f"{meses[i]}: Saldo Anterior={sa:,.2f} ≠ Saldo Atual {meses[i-1]}={prev:,.2f} "
                f"(diff={sa - prev:,.2f})"
            )
    return erros


def check_receitas_menos_despesas(meses, data):
    """RECEITAS - DESPESAS deve bater com a linha 'Receitas - Despesas'."""
    erros = []
    for m in meses:
        rec = v(data, "RECEITAS", m)
        desp = v(data, "DESPESAS", m)
        rd = v(data, "Receitas - Despesas", m)
        if rec is not None and desp is not None and rd is not None:
            calc = rec - desp
            if abs(calc - rd) > 0.02:
                erros.append(
                    f"{m}: {rec:,.2f} - {desp:,.2f} = {calc:,.2f}, "
                    f"mas consta {rd:,.2f} (diff={calc - rd:,.2f})"
                )
    return erros


def check_saldo_balanco(meses, data):
    """Saldo Anterior + (Receitas - Despesas) deve igualar Saldo Atual."""
    erros = []
    for m in meses:
        sa = v(data, "Saldo Anterior", m)
        rd = v(data, "Receitas - Despesas", m)
        saldo = v(data, "Saldo Atual", m)
        if sa is not None and rd is not None and saldo is not None:
            calc = sa + rd
            if abs(calc - saldo) > 0.02:
                erros.append(
                    f"{m}: {sa:,.2f} + {rd:,.2f} = {calc:,.2f}, "
                    f"mas Saldo Atual={saldo:,.2f} (diff={calc - saldo:,.2f})"
                )
    return erros


def check_subtotais_receitas(meses, data):
    """Soma dos subtotais de receita deve igualar RECEITAS."""
    subs_receita = ["Receitas Operacionais", "Receitas Financeiras", "Recuperação de Ativos"]
    erros = []
    for m in meses:
        rec = v(data, "RECEITAS", m)
        if rec is None:
            continue
        calc = sum(v(data, s, m) or 0 for s in subs_receita)
        if abs(calc - rec) > 0.02:
            erros.append(
                f"{m}: subtotais somam {calc:,.2f}, RECEITAS={rec:,.2f} (gap={rec - calc:,.2f})"
            )
    return erros


def check_subtotais_despesas(meses, data):
    """Soma dos subtotais de despesa deve igualar DESPESAS."""
    subs_despesa = ["Despesas com Pessoal", "Despesas Administrativas", "Despesas Financeiras"]
    erros = []
    for m in meses:
        desp = v(data, "DESPESAS", m)
        if desp is None:
            continue
        calc = sum(v(data, s, m) or 0 for s in subs_despesa)
        if abs(calc - desp) > 0.02:
            erros.append(
                f"{m}: subtotais somam {calc:,.2f}, DESPESAS={desp:,.2f} (gap={desp - calc:,.2f})"
            )
    return erros


def check_transitoria_zerada(meses, data):
    """Conta Transitória deve ter saldo líquido zero (receita + despesa = 0)."""
    erros = []
    # Sub-itens podem estar indentados com espaços
    ct_d_key = next((k for k in data if k.strip() == "Movimentação Transitória - Despesa"), None)
    ct_r_key = next((k for k in data if k.strip() == "Movimentação Transitória - Receita"), None)
    if not ct_d_key or not ct_r_key:
        return erros
    for m in meses:
        ct_d = v(data, ct_d_key, m)
        ct_r = v(data, ct_r_key, m)
        if ct_d is not None and ct_r is not None:
            net = ct_d + ct_r
            if abs(net) > 0.02:
                erros.append(f"{m}: Transitória net = {net:,.2f} (deveria ser 0)")
    return erros


def check_dados_faltantes(meses, data):
    """Detecta linhas estruturais (saldo/total/subtotal) com valor ausente."""
    linhas_esperadas = [
        "Saldo Anterior", "RECEITAS", "DESPESAS",
        "Receitas - Despesas", "Saldo Atual",
        "Receitas Operacionais", "Receitas Financeiras",
        "Despesas com Pessoal", "Despesas Administrativas", "Despesas Financeiras",
        "Conta Transitória",
    ]
    erros = []
    for desc_esperada in linhas_esperadas:
        # Encontrar pela descrição (pode estar indentada ou não)
        key = next((k for k in data if k.strip() == desc_esperada), None)
        if key is None:
            continue
        for m in meses:
            val = v(data, key, m)
            if val is None:
                erros.append(f"{m}: '{desc_esperada}' sem valor")
    return erros


def check_itens_recorrentes_ausentes(meses, data):
    """Itens que aparecem na maioria dos meses mas faltam em algum."""
    erros = []
    threshold = max(2, len(meses) * 0.6)  # presente em pelo menos 60% dos meses
    for desc, info in data.items():
        if not isinstance(info, dict) or info.get("tipo") != "item" or desc.startswith("    "):
            continue
        presentes = [m for m in meses if info["vals"].get(m) is not None]
        ausentes = [m for m in meses if info["vals"].get(m) is None]
        if len(presentes) >= threshold and 0 < len(ausentes) <= 3:
            erros.append(
                f"'{desc}': ausente em {', '.join(ausentes)} "
                f"(presente em {len(presentes)}/{len(meses)} meses)"
            )
    return erros


def check_variacao_anomala(meses, data):
    """Variações mês-a-mês superiores a 3x com diferença > R$ 2.000."""
    erros = []
    for desc, info in data.items():
        if not isinstance(info, dict) or info.get("tipo") != "item" or desc.startswith("    "):
            continue
        vals = [(m, info["vals"].get(m)) for m in meses]
        for i in range(1, len(vals)):
            m_prev, v_prev = vals[i - 1]
            m_curr, v_curr = vals[i]
            if not v_prev or not v_curr or v_prev <= 0 or v_curr <= 0:
                continue
            ratio = v_curr / v_prev
            diff = abs(v_curr - v_prev)
            if diff < 2000:
                continue
            if ratio > 3.0:
                erros.append(
                    f"'{desc}': {m_prev}={v_prev:,.2f} → {m_curr}={v_curr:,.2f} "
                    f"({ratio:.1f}x, +{diff:,.2f})"
                )
            elif ratio < 0.33:
                erros.append(
                    f"'{desc}': {m_prev}={v_prev:,.2f} → {m_curr}={v_curr:,.2f} "
                    f"({ratio:.1f}x, -{diff:,.2f})"
                )
    return erros


def check_itens_vs_subtotal(meses, data):
    """Verifica se a soma dos itens sob cada subtotal bate, usando as abas mensais."""
    wb_path = data.get("_wb_path")
    if not wb_path:
        return []
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    erros = []

    for mes in meses:
        if mes not in wb.sheetnames:
            continue
        ws = wb[mes]
        subtotal_atual = None
        subtotal_val = None
        soma_itens = 0.0

        for row in ws.iter_rows(min_row=2, values_only=True):
            desc = row[0] or ""
            val = row[1]
            tipo = row[2] or ""

            if tipo == "subtotal":
                # Fechar subtotal anterior
                if subtotal_atual is not None and subtotal_val is not None:
                    if abs(soma_itens - subtotal_val) > 0.02:
                        erros.append(
                            f"{mes} '{subtotal_atual}': itens somam {soma_itens:,.2f}, "
                            f"subtotal={subtotal_val:,.2f} (gap={subtotal_val - soma_itens:,.2f})"
                        )
                subtotal_atual = desc
                subtotal_val = val
                soma_itens = 0.0
            elif tipo in ("total", "saldo"):
                # Fechar subtotal anterior
                if subtotal_atual is not None and subtotal_val is not None:
                    if abs(soma_itens - subtotal_val) > 0.02:
                        erros.append(
                            f"{mes} '{subtotal_atual}': itens somam {soma_itens:,.2f}, "
                            f"subtotal={subtotal_val:,.2f} (gap={subtotal_val - soma_itens:,.2f})"
                        )
                subtotal_atual = None
            elif tipo == "item" and subtotal_atual and val is not None:
                soma_itens += val

        # Fechar último subtotal pendente
        if subtotal_atual is not None and subtotal_val is not None:
            if abs(soma_itens - subtotal_val) > 0.02:
                erros.append(
                    f"{mes} '{subtotal_atual}': itens somam {soma_itens:,.2f}, "
                    f"subtotal={subtotal_val:,.2f} (gap={subtotal_val - soma_itens:,.2f})"
                )

    return erros


# ── Runner ───────────────────────────────────────────────────────

CHECKS = [
    ("Continuidade do Saldo (Anterior[m] = Atual[m-1])", check_saldo_continuidade),
    ("Receitas - Despesas = linha 'Receitas - Despesas'", check_receitas_menos_despesas),
    ("Saldo Anterior + Rec-Desp = Saldo Atual", check_saldo_balanco),
    ("Subtotais de Receita somam ao total RECEITAS", check_subtotais_receitas),
    ("Subtotais de Despesa somam ao total DESPESAS", check_subtotais_despesas),
    ("Itens somam ao subtotal correspondente", check_itens_vs_subtotal),
    ("Conta Transitória saldo líquido = 0", check_transitoria_zerada),
    ("Dados estruturais ausentes (saldo/total/subtotal)", check_dados_faltantes),
    ("Itens recorrentes ausentes em algum mês", check_itens_recorrentes_ausentes),
    ("Variações anômalas mês-a-mês (>3x e >R$2.000)", check_variacao_anomala),
]


def main():
    parser = argparse.ArgumentParser(description=__doc__.strip().split("\n")[0])
    parser.add_argument("arquivo", nargs="?", default="demonstrativos.xlsx",
                        help="Arquivo Excel a analisar (default: demonstrativos.xlsx)")
    args = parser.parse_args()

    if not Path(args.arquivo).exists():
        print(f"❌ Arquivo '{args.arquivo}' não encontrado.")
        sys.exit(1)

    meses, data = carregar_evolucao(args.arquivo)
    print(f"📊 Analisando {args.arquivo} — {len(meses)} meses ({meses[0]} a {meses[-1]})")
    print(f"   {len(data) - 1} linhas na aba Evolução\n")

    total_erros = 0
    for titulo, check_fn in CHECKS:
        erros = check_fn(meses, data)
        if erros:
            print(f"❌ {titulo}")
            for e in erros:
                print(f"   • {e}")
            print()
        else:
            print(f"✅ {titulo}")
        total_erros += len(erros)

    print()
    if total_erros == 0:
        print("🎉 Nenhum problema encontrado!")
    else:
        print(f"⚠️  {total_erros} problema(s) encontrado(s).")


if __name__ == "__main__":
    main()
