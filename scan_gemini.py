"""
Extrai dados de demonstrativo financeiro usando Gemini 2.5 Flash.

Uso:
    pixi run python scan_gemini.py                    # processa todas as imagens de ./imagens/
    pixi run python scan_gemini.py 2026-03.jpg        # processa imagens específicas
    pixi run python scan_gemini.py --pasta fotos/     # processa imagens de outra pasta

Variáveis de ambiente:
    GEMINI_API_KEY  — chave da Google AI Studio (https://aistudio.google.com/apikey)

Resultados são cacheados em .cache_gemini/ por hash do arquivo, então re-rodar é grátis.
Saída: planilha demonstrativos.xlsx com evolução mensal e gráficos.
"""

import argparse
import base64
from collections import Counter
import hashlib
import json
import os
import re
import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import Outline
import plotly.graph_objects as go
from google import genai
from google.genai import types

# ── Configuração ─────────────────────────────────────────────────
MODELO = "gemini-2.5-flash"
MODELO_FALLBACK = "gemini-3.1-flash-lite-preview"  # 500 RPD vs 20 RPD do Flash
CACHE_DIR = Path(".cache_gemini")
PASTA_IMAGENS = Path("imagens")
EXTENSOES_IMG = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif", ".pdf"}
QUOTA_DIARIA = 20  # limite free tier RPD para gemini-2.5-flash
USAGE_FILE = CACHE_DIR / "_usage.json"

PROMPT = """Você está olhando para um demonstrativo financeiro mensal de um CONDOMÍNIO brasileiro
(Demonstrativo de Receitas e Despesas). Sua tarefa: extrair TODAS as linhas que contêm um valor
monetário, preservando a ordem em que aparecem na página.

⚠️ CRÍTICO — ALINHAMENTO DESCRIÇÃO↔VALOR:
  - Cada linha horízontal da imagem tem UMA descrição (à esquerda) e UM valor (à direita).
  - NUNCA pule uma linha. Se você vê um valor mas não consegue ler bem a descrição,
    inclua mesmo assim com a melhor leitura possível — nunca puxe o valor da próxima linha.
  - Linhas pequenas/curtas (ex: "Multas e Juros" com valor 63,57) são fáceis de pular: NO ATTENTION,
    leia linha-por-linha de cima para baixo.
  - Conte as linhas com valor antes de retornar e confirme que descrições e valores têm o mesmo número.

Para cada linha, retorne:
- "descricao": texto descritivo da linha, limpo (sem pontos de líder "....", sem referências de página tipo "9 a 20")
- "valor": número decimal usando PONTO como separador decimal (ex: 1308.10, -39664.48). Negativos com sinal.
- "tipo": classifique como uma destas categorias:
    * "saldo"     — Saldo Anterior, Saldo Atual
    * "total"     — RECEITAS, DESPESAS, Receitas - Despesas (totais em CAIXA ALTA ou de seção)
    * "subtotal"  — Receitas Operacionais, Receitas Financeiras, Despesas com Pessoal,
                    Despesas Administrativas, Despesas Financeiras, Conta Transitória
    * "item"      — todas as demais linhas individuais

Regras adicionais:
- Inclua linhas com valor 0,00 (Conta Transitória pode ter zero).
- Preserve nomes próprios e números de NF/protocolo (ex: "NF 153", "NFCe 60624").
- Use grafia portuguesa correta (Síndico, Refeição, etc).

Retorne APENAS um array JSON, sem texto extra.
"""

# ── Normalização de descrições via LLM (variantes OCR → forma canônica) ───
NORMALIZACAO_CACHE = CACHE_DIR / "_normalizacao.json"

# Ordem dos tipos para agrupar na aba Evolução
_TIPO_ORDEM = {"saldo": 0, "total": 1, "subtotal": 2, "item": 3}

PROMPT_NORMALIZACAO = """Você recebe uma lista de descrições extraídas por OCR de demonstrativos financeiros
mensais de um condomínio. Muitas são variantes da mesma coisa (acentuação, pontuação, espaçamento).

Sua tarefa: identificar grupos de descrições que se referem ao MESMO item e escolher a forma
canônica (mais correta/completa) para cada grupo.

Regras:
- Só agrupe descrições que realmente são o mesmo item (ex: "Honorario Desterro" e "Honorário Desterro").
- NÃO agrupe itens diferentes que apenas têm prefixo parecido mas NF/parcela diferente
  (ex: "Manutenção Elevadores - Elevacon NF 12255" e "NF 12495" são pagamentos distintos).
- NÃO agrupe "Seguros" (subtotal) com "Seguro - Axa" (item individual) — tipos diferentes.
- A forma canônica deve usar acentuação correta do português.
- NÃO expanda abreviações (mantenha "s/", "c/", "Manut.", "Desp." como estão).
- NÃO adicione hifens, preposições ou mude formatação além de corrigir acentuação e espaçamento.
- Descrições que não têm variantes não precisam aparecer no resultado.

Retorne APENAS um objeto JSON onde cada chave é a descrição original e o valor é a forma canônica.
Inclua SOMENTE as descrições que precisam ser renomeadas (variantes → canônica).
Descrições que já estão na forma correta e não têm variantes devem ser OMITIDAS.

Exemplo de saída:
{"Honorario Desterro": "Honorário Desterro", "Manut.Contra Incendio": "Manut. Contra Incêndio"}
"""


def _carregar_normalizacao() -> tuple[dict[str, str], set[str]]:
    """Carrega mapa de normalização e conjunto de descrições já vistas."""
    if NORMALIZACAO_CACHE.exists():
        data = json.loads(NORMALIZACAO_CACHE.read_text(encoding="utf-8"))
        if isinstance(data, dict) and "mapa" in data:
            return data["mapa"], set(data.get("vistas", []))
        # migrar formato antigo (só mapa)
        return data, set(data.keys()) | set(data.values())
    return {}, set()


def _salvar_normalizacao(mapa: dict[str, str], vistas: set[str]):
    """Salva mapa + descrições vistas no cache."""
    CACHE_DIR.mkdir(exist_ok=True)
    data = {"mapa": mapa, "vistas": sorted(vistas)}
    NORMALIZACAO_CACHE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def _gerar_normalizacao(todas_desc: list[str], client: genai.Client, modelo: str) -> dict[str, str]:
    """Gera/atualiza mapa de normalização via LLM. Só chama a API se há descrições novas."""
    mapa, vistas = _carregar_normalizacao()

    novas = [d for d in todas_desc if d not in vistas]

    if not novas:
        return mapa

    print(f"\n🔤 {len(novas)} descrições novas para normalizar...")
    # Enviar TODAS as descrições (não só as novas) para contexto completo
    lista_txt = "\n".join(f"- {d}" for d in sorted(todas_desc))
    prompt_completo = PROMPT_NORMALIZACAO + "\n\nDescrições:\n" + lista_txt

    cfg_kwargs = dict(response_mime_type="application/json", temperature=0)
    if "lite" not in modelo.lower():
        cfg_kwargs["thinking_config"] = types.ThinkingConfig(thinking_budget=8000)

    resp = client.models.generate_content(
        model=modelo,
        contents=[prompt_completo],
        config=types.GenerateContentConfig(**cfg_kwargs),
    )
    _registrar_chamada()

    novo_mapa = json.loads(resp.text)
    # Novas regras do LLM, mas manter edições manuais existentes (não sobrescrever)
    novo_mapa.update(mapa)
    mapa = novo_mapa

    # Marcar TODAS as descrições atuais como vistas
    vistas.update(todas_desc)

    _salvar_normalizacao(mapa, vistas)
    print(f"   💾 Mapa de normalização salvo ({len(mapa)} regras)")
    return mapa


# Mapa global preenchido em main()
_MAPA_NORMALIZACAO: dict[str, str] = {}
_MAPA_AGRUPAMENTO: dict[str, list[str]] = {}  # {parent: [children]}


def _normalizar(desc: str) -> str:
    return _MAPA_NORMALIZACAO.get(desc, desc)


def _normalizar_todos(todos: dict[str, list[dict]]) -> dict[str, list[dict]]:
    """Retorna cópia de `todos` com descrições normalizadas."""
    out = {}
    for mes, rows in todos.items():
        out[mes] = [{**r, "descricao": _normalizar(r.get("descricao", ""))} for r in rows]
    return out


_RE_NF = re.compile(r'\s+[-–]?\s*NF[Cc]?[Ee]?\s*\d+.*$')
_RE_PARCELA = re.compile(r'\s*-?\s*\d+/\d+\s*$')


def _agregar_descricao(desc: str) -> str:
    """Strip NF/NFCe numbers and installment suffixes for Evolução aggregation."""
    base = _RE_NF.sub('', desc)
    base = _RE_PARCELA.sub('', base)
    return base.rstrip(' -')


# ── Agrupamento de itens relacionados via LLM ─────────────────────
AGRUPAMENTO_CACHE = CACHE_DIR / "_agrupamento.json"

PROMPT_AGRUPAMENTO = """Você recebe uma lista de descrições de itens financeiros (já agregados, sem números de NF)
de demonstrativos mensais de um condomínio.

Sua tarefa: identificar grupos de itens relacionados que deveriam ser colapsados sob um nome-pai comum.

Regras:
- Só agrupe itens que são claramente variantes ou subtipos da mesma categoria.
- O nome-pai deve ser curto e genérico (ex: "Manutenção Elevadores" para toda manutenção de elevadores).
- Um grupo precisa ter pelo menos 2 itens.
- Itens que não têm itens relacionados NÃO devem aparecer.
- Itens recorrentes que aparecem todo mês sozinhos (ex: "Energia Elétrica", "Telefone") NÃO devem ser agrupados.
- Se um item genérico já existe (ex: "Férias") e há variantes ("Férias Albanisia", "Férias Lucas"),
  inclua o genérico como filho também.

Retorne APENAS um objeto JSON onde cada chave é o nome do grupo (pai) e o valor é um array
com os nomes dos itens filhos. Omita grupos com menos de 2 itens.

Exemplo:
{"Manutenção Elevadores": ["Manutenção Elevadores - Elevacon", "Serviços Manut. Elevadores"]}
"""


def _carregar_agrupamento() -> tuple[dict[str, list[str]], set[str]]:
    if AGRUPAMENTO_CACHE.exists():
        data = json.loads(AGRUPAMENTO_CACHE.read_text(encoding="utf-8"))
        if isinstance(data, dict) and "mapa" in data:
            return data["mapa"], set(data.get("vistas", []))
    return {}, set()


def _salvar_agrupamento(mapa: dict[str, list[str]], vistas: set[str]):
    CACHE_DIR.mkdir(exist_ok=True)
    data = {"mapa": mapa, "vistas": sorted(vistas)}
    AGRUPAMENTO_CACHE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def _gerar_agrupamento(todas_desc: list[str], client: genai.Client, modelo: str) -> dict[str, list[str]]:
    """Gera/atualiza mapa de agrupamento via LLM. Só chama a API se há descrições novas."""
    mapa, vistas = _carregar_agrupamento()

    novas = [d for d in todas_desc if d not in vistas]
    if not novas:
        return mapa

    print(f"\n📦 {len(novas)} descrições novas para agrupar...")
    lista_txt = "\n".join(f"- {d}" for d in sorted(todas_desc))
    prompt_completo = PROMPT_AGRUPAMENTO + "\nDescrições:\n" + lista_txt

    cfg_kwargs = dict(response_mime_type="application/json", temperature=0)
    if "lite" not in modelo.lower():
        cfg_kwargs["thinking_config"] = types.ThinkingConfig(thinking_budget=8000)

    resp = client.models.generate_content(
        model=modelo,
        contents=[prompt_completo],
        config=types.GenerateContentConfig(**cfg_kwargs),
    )
    _registrar_chamada()

    novo_mapa = json.loads(resp.text)
    # Existentes têm prioridade sobre LLM (edições manuais)
    novo_mapa.update(mapa)
    mapa = novo_mapa

    vistas.update(todas_desc)
    _salvar_agrupamento(mapa, vistas)
    print(f"   💾 Mapa de agrupamento salvo ({len(mapa)} grupos)")
    return mapa


def _ler_uso() -> dict:
    """Lê contador de uso diário. Reseta se o dia mudou."""
    if USAGE_FILE.exists():
        data = json.loads(USAGE_FILE.read_text(encoding="utf-8"))
        if data.get("data") == str(date.today()):
            return data
    return {"data": str(date.today()), "chamadas": 0}


def _registrar_chamada():
    """Incrementa contador de uso diário."""
    uso = _ler_uso()
    uso["chamadas"] += 1
    CACHE_DIR.mkdir(exist_ok=True)
    USAGE_FILE.write_text(json.dumps(uso), encoding="utf-8")


def _verificar_quota(n_pendentes: int):
    """Verifica se há quota suficiente antes de chamar a API."""
    uso = _ler_uso()
    restante = QUOTA_DIARIA - uso["chamadas"]
    print(f"   📊 Quota: {uso['chamadas']}/{QUOTA_DIARIA} usadas hoje, {restante} restantes")
    if restante <= 0:
        raise RuntimeError(
            f"⛔ Quota diária esgotada ({uso['chamadas']}/{QUOTA_DIARIA}). "
            "Aguarde reset (meia-noite PT). Imagens já cacheadas continuam funcionando."
        )
    if n_pendentes > restante:
        print(f"   ⚠️  {n_pendentes} imagens novas mas só {restante} chamadas restantes — "
              f"processando apenas as primeiras {restante}")


def extrair_via_gemini(image_path: Path, client: genai.Client, modelo: str = MODELO) -> list[dict]:
    """Extrai linhas do demonstrativo. Cacheia resultado por hash do arquivo."""
    img_bytes = image_path.read_bytes()
    img_hash = hashlib.sha256(img_bytes).hexdigest()[:16]
    cache_file = CACHE_DIR / f"{image_path.stem}_{img_hash}.json"

    if cache_file.exists():
        print(f"   📦 cache hit ({cache_file.name})")
        return json.loads(cache_file.read_text(encoding="utf-8"))

    print(f"   🌐 chamando {modelo}...")
    uso = _ler_uso()
    if uso["chamadas"] >= QUOTA_DIARIA:
        raise RuntimeError(
            f"⛔ Quota diária esgotada ({uso['chamadas']}/{QUOTA_DIARIA}). "
            "Aguarde reset (meia-noite PT)."
        )
    mime = "image/jpeg" if image_path.suffix.lower() in (".jpg", ".jpeg") else "image/png"
    if image_path.suffix.lower() == ".pdf":
        mime = "application/pdf"

    # Thinking config só para modelos que suportam (não-Lite)
    cfg_kwargs = dict(
        response_mime_type="application/json",
        temperature=0,
    )
    if "lite" not in modelo.lower():
        cfg_kwargs["thinking_config"] = types.ThinkingConfig(thinking_budget=8000)

    try:
        resp = client.models.generate_content(
            model=modelo,
            contents=[
                types.Part.from_bytes(data=img_bytes, mime_type=mime),
                PROMPT,
            ],
            config=types.GenerateContentConfig(**cfg_kwargs),
        )
        _registrar_chamada()
    except Exception as e:
        _registrar_chamada()  # falhas também contam na quota
        msg = str(e)
        if "429" in msg or "RESOURCE_EXHAUSTED" in msg:
            raise RuntimeError(
                "⛔ Quota diária esgotada. Aguarde reset (meia-noite PT). "
                "Imagens já cacheadas continuam funcionando."
            ) from e
        if "503" in msg or "UNAVAILABLE" in msg:
            raise RuntimeError(
                "⛔ Modelo sobrecarregado (503 também gasta quota!). "
                "Tente novamente mais tarde."
            ) from e
        raise

    rows = json.loads(resp.text)
    CACHE_DIR.mkdir(exist_ok=True)
    cache_file.write_text(
        json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"   💾 cache salvo em {cache_file.name}")
    return rows


# ── Estilos Excel ────────────────────────────────────────────────
COR_TIPO = {
    "saldo":    PatternFill("solid", fgColor="FFE699"),  # amarelo
    "total":    PatternFill("solid", fgColor="B4C7E7"),  # azul
    "subtotal": PatternFill("solid", fgColor="D9E1F2"),  # azul claro
    "item":     None,
}
FONTE_DESTAQUE = Font(bold=True)


def escrever_aba(ws, rows: list[dict], titulo: str):
    ws.title = titulo[:31]  # limite Excel
    ws.append(["Descrição", "Valor", "Tipo"])
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="305496")
        c.font = Font(bold=True, color="FFFFFF")

    for row in rows:
        ws.append([row.get("descricao", ""), row.get("valor"), row.get("tipo", "item")])
        excel_row = ws.max_row
        tipo = row.get("tipo", "item")
        fill = COR_TIPO.get(tipo)
        if fill:
            for cell in ws[excel_row]:
                cell.fill = fill
        if tipo in ("saldo", "total", "subtotal"):
            for cell in ws[excel_row]:
                cell.font = FONTE_DESTAQUE
        # formato número BR
        ws.cell(row=excel_row, column=2).number_format = '#,##0.00;-#,##0.00'

    ws.column_dimensions["A"].width = 65
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = "A2"


def descobrir_imagens(pasta: Path) -> list[Path]:
    """Encontra imagens/PDFs na pasta, ordenadas por nome (espera-se YYYY-MM)."""
    imgs = [f for f in sorted(pasta.iterdir()) if f.suffix.lower() in EXTENSOES_IMG]
    return imgs


def construir_evolucao(wb: openpyxl.Workbook, todos: dict[str, list[dict]]):
    """Cria aba 'Evolução' com pivot mensal + gráficos embutidos."""
    todos = _normalizar_todos(todos)
    meses = list(todos.keys())  # já ordenados
    if not meses:
        return

    # ── Determinar seção de cada descrição a partir da ordem do documento ──
    # Seções do demonstrativo na ordem natural:
    #   0: Saldo Anterior
    #   1: RECEITAS (subtotais + itens de receita)
    #   2: DESPESAS (subtotais + itens de despesa)
    #   3: CONTA TRANSITÓRIA
    #   4: Receitas - Despesas
    #   5: Saldo Atual
    SECAO_HEADERS = {
        "Saldo Anterior": 0,
        "RECEITAS": 1,
        "DESPESAS": 2,
        "Conta Transitória": 3, "CONTA TRANSITÓRIA": 3,
        "Receitas - Despesas": 4,
        "Saldo Atual": 5,
    }

    # Determinar seção e agregar itens (strip NF/parcela, somar valores)
    secao_de = {}  # agg_desc → seção (int)
    tipo_de = {}   # agg_desc → tipo
    pivot = {}     # {agg_desc: {mes: valor}}
    # Rastrear originais: {agg_desc: {mes: [(desc_original, valor), ...]}}
    originais = {}
    subtotal_de = {}  # agg_desc(item) → subtotal_desc
    todas_desc = []
    desc_set = set()

    for mes in meses:
        secao_atual = -1
        subtotal_atual = None
        for r in todos[mes]:
            d = r.get("descricao", "")
            if not d:
                continue
            # Atualizar seção corrente
            if d in SECAO_HEADERS:
                secao_atual = SECAO_HEADERS[d]

            tipo = r.get("tipo", "item")
            # Só agregar itens — totais/subtotais/saldos ficam como estão
            agg = _normalizar(_agregar_descricao(d)) if tipo == "item" else d

            # Rastrear hierarquia subtotal → itens
            if tipo == "subtotal":
                subtotal_atual = agg
            elif tipo in ("total", "saldo"):
                subtotal_atual = None
            elif tipo == "item" and subtotal_atual and agg not in subtotal_de:
                subtotal_de[agg] = subtotal_atual

            if agg not in secao_de:
                secao_de[agg] = max(secao_atual, 0)
            if agg not in tipo_de:
                tipo_de[agg] = tipo
            if agg not in desc_set:
                todas_desc.append(agg)
                desc_set.add(agg)

            # Somar valores no pivot (múltiplas NFs → mesmo mês)
            val = r.get("valor")
            if val is not None:
                bucket = pivot.setdefault(agg, {})
                bucket[mes] = (bucket.get(mes) or 0) + val
                # Guardar originais quando a descrição difere da agregada
                if d != agg:
                    originais.setdefault(agg, {}).setdefault(mes, []).append(
                        (d, val)
                    )

    # Ordenar: por seção do documento, depois por tipo, depois alfabeticamente
    todas_desc.sort(key=lambda d: (
        secao_de.get(d, 99),
        _TIPO_ORDEM.get(tipo_de.get(d, "item"), 3),
        d.lower(),
    ))

    # ── Agrupamento hierárquico ──
    # Mapear child → parent para itens que existem nos dados
    # Pular grupos cujo nome já é um subtotal do documento (hierarquia já existe)
    subtotais_existentes = {d for d in desc_set if tipo_de.get(d) == "subtotal"}
    grupo_de = {}   # child_desc → parent_name
    filhos_de = {}  # parent_name → [child_desc, ...]
    for parent, children in _MAPA_AGRUPAMENTO.items():
        if parent in subtotais_existentes:
            continue
        existentes = [c for c in children if c in desc_set and c not in subtotais_existentes]
        if len(existentes) >= 2:
            # Agrupar apenas filhos na mesma seção (majoritária)
            sec_counts = Counter(secao_de.get(c, 2) for c in existentes)
            secao_maj = sec_counts.most_common(1)[0][0]
            mesma_secao = [c for c in existentes if secao_de.get(c, 2) == secao_maj]
            if len(mesma_secao) >= 2:
                filhos_de[parent] = sorted(mesma_secao, key=str.lower)
                for c in mesma_secao:
                    grupo_de[c] = parent

    # Para cada pai, coletar linhas originais (pré-agregação) dos filhos
    originais_grupo = {}  # {parent: {orig_desc: {mes: valor}}}
    for parent, children in filhos_de.items():
        raw = {}
        for child in children:
            child_orig = originais.get(child, {})
            if child_orig:
                # Filho tinha NFs agregadas – usar as linhas originais
                for mes, itens in child_orig.items():
                    for orig_desc, val in itens:
                        raw.setdefault(orig_desc, {})[mes] = \
                            (raw.get(orig_desc, {}).get(mes) or 0) + val
            else:
                # Filho sem agregação – ele próprio é o item original
                for mes, val in pivot.get(child, {}).items():
                    raw.setdefault(child, {})[mes] = val
        originais_grupo[parent] = raw

    # Calcular pivot para pais (soma dos filhos) e herdar seção
    for parent, children in filhos_de.items():
        parent_vals = {}
        for child in children:
            for mes in meses:
                v = pivot.get(child, {}).get(mes)
                if v is not None:
                    parent_vals[mes] = (parent_vals.get(mes) or 0) + v
        pivot[parent] = parent_vals
        if parent not in secao_de:
            # Usar a seção mais comum entre os filhos (voto majoritário)
            sec_counts = Counter(secao_de.get(c, 2) for c in children)
            secao_de[parent] = sec_counts.most_common(1)[0][0]
        tipo_de[parent] = "item"
        # Herdar subtotal dos filhos
        if parent not in subtotal_de:
            child_subs = [subtotal_de[c] for c in children if c in subtotal_de]
            if child_subs:
                subtotal_de[parent] = Counter(child_subs).most_common(1)[0][0]

    # Reconstruir lista: remover filhos e inserir o pai na posição correta
    todas_desc_final = [d for d in todas_desc if d not in grupo_de]
    # Inserir pais na lista e re-ordenar pelo mesmo critério
    for parent in filhos_de:
        todas_desc_final.append(parent)
    todas_desc_final.sort(key=lambda d: (
        secao_de.get(d, 99),
        _TIPO_ORDEM.get(tipo_de.get(d, "item"), 3),
        d.lower(),
    ))

    # Escrever aba
    ws = wb.create_sheet("Evolução", 0)
    ws.sheet_properties.outlinePr = Outline(summaryBelow=False)
    # Cabeçalho
    ws.append(["Tipo", "Descrição"] + meses)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="305496")
    ws.freeze_panes = "C2"

    FILL_SUBITEM = PatternFill("solid", fgColor="F2F2F2")  # cinza claro
    FONT_SUBITEM = Font(color="666666")

    def _escrever_linha(desc, valores, tipo, outline_level=0):
        indent = "    " * outline_level
        row_data = [tipo, indent + desc]
        for mes in meses:
            row_data.append(valores.get(mes))
        ws.append(row_data)

        excel_row = ws.max_row
        fill = COR_TIPO.get(tipo)
        if outline_level >= 1:
            fill = FILL_SUBITEM
        if fill:
            for cell in ws[excel_row]:
                cell.fill = fill
        if tipo in ("saldo", "total", "subtotal"):
            for cell in ws[excel_row]:
                cell.font = FONTE_DESTAQUE
        elif outline_level >= 1:
            for cell in ws[excel_row]:
                cell.font = FONT_SUBITEM

        for i, mes in enumerate(meses):
            col = 3 + i
            cell = ws.cell(row=excel_row, column=col)
            cell.number_format = '#,##0.00;[Red]-#,##0.00'

        if outline_level > 0:
            ws.row_dimensions[excel_row].outlineLevel = outline_level
            ws.row_dimensions[excel_row].hidden = True

    # Construir mapa subtotal → [filhos na ordem da lista final]
    filhos_de_subtotal = {}
    itens_sob_subtotal = set()
    for d in todas_desc_final:
        sub = subtotal_de.get(d)
        tipo = tipo_de.get(d, "item")
        if sub and tipo == "item":
            filhos_de_subtotal.setdefault(sub, []).append(d)
            itens_sob_subtotal.add(d)

    for desc in todas_desc_final:
        if desc in itens_sob_subtotal:
            continue  # será escrito como filho do subtotal

        tipo = tipo_de.get(desc, "item")
        _escrever_linha(desc, pivot.get(desc, {}), tipo)

        # Se é um subtotal, escrever seus itens (sempre visíveis)
        if desc in filhos_de_subtotal:
            for child in filhos_de_subtotal[desc]:
                _escrever_linha(child, pivot.get(child, {}), "item")

                # Se o filho é um grupo LLM, escrever linhas originais colapsadas
                if child in originais_grupo:
                    for orig in sorted(originais_grupo[child], key=str.lower):
                        _escrever_linha(
                            orig, originais_grupo[child][orig],
                            "item", outline_level=1,
                        )

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 55
    for i in range(len(meses)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 14

    if len(meses) < 2:
        return  # gráficos só fazem sentido com 2+ meses


def gerar_html(todos: dict[str, list[dict]], img_paths: dict[str, Path], output: Path):
    """Gera página HTML com gráficos interativos Plotly + visualizador de imagens."""
    todos = _normalizar_todos(todos)
    meses = list(todos.keys())
    if len(meses) < 2:
        print("   ⚠️  Menos de 2 meses — HTML não gerado (precisa de evolução).")
        return

    # Montar pivot: {desc: {mes: valor}}
    pivot = {}
    for mes in meses:
        for r in todos[mes]:
            d = r.get("descricao", "")
            if d:
                pivot.setdefault(d, {})[mes] = r.get("valor")

    def serie(nome):
        return [pivot.get(nome, {}).get(m) for m in meses]

    COMMON = dict(template="plotly_white", hovermode="x unified",
                  margin=dict(l=60, r=30, t=50, b=40), dragmode=False)

    # ── Chart 1: Receitas × Despesas × Saldo Atual (line) ──
    fig1 = go.Figure()
    for nome, cor in [("RECEITAS", "#2ecc71"), ("DESPESAS", "#e74c3c"), ("Saldo Atual", "#3498db")]:
        fig1.add_trace(go.Scatter(
            x=meses, y=serie(nome), name=nome, mode="lines+markers",
            line=dict(color=cor, width=3),
            hovertemplate="R$ %{y:,.2f}<extra>" + nome + "</extra>",
        ))
    fig1.update_layout(
        title="Receitas × Despesas × Saldo Atual", height=420,
        yaxis=dict(title="R$", tickformat=",.0f"),
        xaxis=dict(showspikes=True, spikemode="across", spikesnap="cursor",
                   spikethickness=1, spikecolor="#888", spikedash="dot"),
        legend=dict(orientation="h", yanchor="bottom", y=-0.25),
        **COMMON,
    )

    # ── Chart 2: Composição das Receitas (stacked bar) ──
    fig2 = go.Figure()
    for nome, cor in [("Receitas Operacionais", "#27ae60"), ("Receitas Financeiras", "#82e0aa")]:
        fig2.add_trace(go.Bar(
            x=meses, y=serie(nome), name=nome, marker_color=cor,
            hovertemplate="R$ %{y:,.2f}<extra>" + nome + "</extra>",
        ))
    fig2.update_layout(
        title="Composição das Receitas", barmode="stack", height=380,
        yaxis=dict(title="R$", tickformat=",.0f"),
        legend=dict(orientation="h", yanchor="bottom", y=-0.25),
        **COMMON,
    )

    # ── Chart 3: Composição das Despesas (stacked bar) ──
    fig3 = go.Figure()
    cores_desp = [("Despesas com Pessoal", "#e74c3c"), ("Despesas Administrativas", "#e67e22"), ("Despesas Financeiras", "#9b59b6")]
    for nome, cor in cores_desp:
        fig3.add_trace(go.Bar(
            x=meses, y=serie(nome), name=nome, marker_color=cor,
            hovertemplate="R$ %{y:,.2f}<extra>" + nome + "</extra>",
        ))
    fig3.update_layout(
        title="Composição das Despesas", barmode="stack", height=380,
        yaxis=dict(title="R$", tickformat=",.0f"),
        legend=dict(orientation="h", yanchor="bottom", y=-0.25),
        **COMMON,
    )

    # ── Montar HTML completo ──
    pcfg = {"scrollZoom": False, "displayModeBar": False, "responsive": True}
    div1 = fig1.to_html(include_plotlyjs="cdn", full_html=False, div_id="chart1", config=pcfg)
    div2 = fig2.to_html(include_plotlyjs=False, full_html=False, div_id="chart2", config=pcfg)
    div3 = fig3.to_html(include_plotlyjs=False, full_html=False, div_id="chart3", config=pcfg)

    # ── Encode imagens como base64 ──
    img_data_js = {}
    for mes in meses:
        p = img_paths.get(mes)
        if p and p.exists():
            b64 = base64.b64encode(p.read_bytes()).decode("ascii")
            suffix = p.suffix.lower()
            mime = "image/jpeg" if suffix in (".jpg", ".jpeg") else "image/png"
            img_data_js[mes] = f"data:{mime};base64,{b64}"

    options_html = "\n".join(
        f'        <option value="{m}">{m}</option>' for m in meses
    )
    img_map_js = json.dumps(img_data_js)

    # ── Resumo do último mês ──
    ultimo = meses[-1]
    def val_ultimo(nome):
        v = pivot.get(nome, {}).get(ultimo)
        if v is None:
            return "—"
        return f"R$ {v:,.2f}"
    resumo_saldo = val_ultimo("Saldo Atual")
    resumo_receitas = val_ultimo("RECEITAS")
    resumo_despesas = val_ultimo("DESPESAS")

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1, maximum-scale=1, user-scalable=no">
<title>Condomínio Beverly Boulevard — Evolução Financeira</title>
<style>
  body {{ font-family: system-ui, sans-serif; background: #f8f9fa; margin: 0; padding: 20px; }}
  h1 {{ text-align: center; color: #2c3e50; margin-bottom: 30px; }}
  .chart {{ max-width: 1000px; margin: 0 auto 30px; background: white;
            border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); padding: 10px;
            min-height: 380px; overflow-x: auto; }}
  .viewer {{ max-width: 1000px; margin: 0 auto 30px; background: white;
             border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); padding: 20px; }}
  .viewer h2 {{ margin: 0 0 15px; color: #2c3e50; }}
  .viewer select {{ font-size: 16px; padding: 6px 12px; border-radius: 4px;
                    border: 1px solid #ccc; margin-bottom: 15px; }}
  .viewer img {{ max-width: 100%; border: 1px solid #e0e0e0; border-radius: 4px; }}
  .viewer .placeholder {{ color: #999; font-style: italic; }}
  .summary {{ max-width: 1000px; margin: 0 auto 30px; display: flex; gap: 15px; flex-wrap: wrap; }}
  .summary .card {{ flex: 1; min-width: 200px; background: white; border-radius: 8px;
                    box-shadow: 0 2px 8px rgba(0,0,0,0.1); padding: 20px; text-align: center; }}
  .summary .card .label {{ font-size: 13px; color: #888; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 6px; }}
  .summary .card .value {{ font-size: 26px; font-weight: 700; }}
  .summary .card .month {{ font-size: 12px; color: #aaa; margin-top: 4px; }}
  .summary .card.receitas .value {{ color: #2ecc71; }}
  .summary .card.despesas .value {{ color: #e74c3c; }}
  .summary .card.saldo .value {{ color: #3498db; }}
</style>
</head>
<body>
<h1>Condomínio Beverly Boulevard — Evolução Financeira</h1>
<div class="summary">
  <div class="card receitas"><div class="label">Receitas</div><div class="value">{resumo_receitas}</div><div class="month">{ultimo}</div></div>
  <div class="card despesas"><div class="label">Despesas</div><div class="value">{resumo_despesas}</div><div class="month">{ultimo}</div></div>
  <div class="card saldo"><div class="label">Saldo Atual</div><div class="value">{resumo_saldo}</div><div class="month">{ultimo}</div></div>
</div>
<div class="chart">{div1}</div>
<div class="chart">{div2}</div>
<div class="chart">{div3}</div>
<div class="viewer">
  <h2>Demonstrativo Original</h2>
  <select id="mesSelect" onchange="showImage()">
    <option value="">Selecione o mês…</option>
{options_html}
  </select>
  <div id="imgContainer"><p class="placeholder">Selecione um mês acima para visualizar o demonstrativo.</p></div>
</div>
<script>
var imgMap = {img_map_js};
function showImage() {{
  var mes = document.getElementById('mesSelect').value;
  var container = document.getElementById('imgContainer');
  if (mes && imgMap[mes]) {{
    container.innerHTML = '<img src="' + imgMap[mes] + '" alt="Demonstrativo ' + mes + '">';
  }} else {{
    container.innerHTML = '<p class="placeholder">Imagem não disponível para este mês.</p>';
  }}
}}
</script>
</body>
</html>"""

    output.write_text(html, encoding="utf-8")
    print(f"   📊 HTML salvo: {output}")


def main():
    parser = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    parser.add_argument("imagens", nargs="*", help="Imagens específicas (opcional)")
    parser.add_argument("--pasta", type=Path, default=PASTA_IMAGENS,
                        help=f"Pasta com imagens (default: {PASTA_IMAGENS}/)")
    parser.add_argument("--saida", default="demonstrativos.xlsx", help="Excel de saída")
    parser.add_argument("--modelo", default=MODELO, help=f"Modelo Gemini (default: {MODELO}, fallback: {MODELO_FALLBACK})")
    parser.add_argument("--no-cache", action="store_true", help="Ignora cache")
    args = parser.parse_args()

    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        print("❌ Defina a variável GEMINI_API_KEY (https://aistudio.google.com/apikey)")
        sys.exit(1)

    client = genai.Client(api_key=api_key)

    if args.no_cache and CACHE_DIR.exists():
        for f in CACHE_DIR.glob("*.json"):
            f.unlink()

    # Determinar imagens: args explícitos ou auto-scan da pasta
    if args.imagens:
        img_paths = [Path(p) for p in args.imagens]
    else:
        if not args.pasta.exists():
            args.pasta.mkdir(parents=True)
            print(f"📁 Pasta '{args.pasta}/' criada. Coloque as imagens lá e rode novamente.")
            sys.exit(0)
        img_paths = descobrir_imagens(args.pasta)
        if not img_paths:
            print(f"📁 Nenhuma imagem encontrada em '{args.pasta}/'. "
                  f"Extensões aceitas: {', '.join(sorted(EXTENSOES_IMG))}")
            sys.exit(0)

    print(f"📋 {len(img_paths)} imagem(ns) para processar")

    # Contar quantas precisam de API (sem cache)
    n_sem_cache = 0
    for path in sorted(img_paths):
        img_hash = hashlib.sha256(path.read_bytes()).hexdigest()[:16]
        cache_file = CACHE_DIR / f"{path.stem}_{img_hash}.json"
        if not cache_file.exists():
            n_sem_cache += 1
    if n_sem_cache > 0:
        _verificar_quota(n_sem_cache)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    todos = {}  # {mes_label: rows}

    for path in sorted(img_paths):
        if not path.exists():
            print(f"⚠️  pulando {path} (não existe)")
            continue
        print(f"\n📄 {path.name}")
        try:
            rows = extrair_via_gemini(path, client, args.modelo)
        except Exception as e:
            print(f"   ❌ erro: {e}")
            continue
        print(f"   ✅ {len(rows)} linhas extraídas")

        mes_label = path.stem  # ex: "2025-12", "2026-02"
        todos[mes_label] = rows

        ws = wb.create_sheet()
        escrever_aba(ws, rows, mes_label)

    # Normalização de descrições via LLM (cached)
    global _MAPA_NORMALIZACAO, _MAPA_AGRUPAMENTO
    desc_raw = {r.get("descricao", "") for rows in todos.values() for r in rows} - {""}
    desc_agg = {_agregar_descricao(d) for d in desc_raw} - {""}
    desc_norm = {_normalizar(d) for d in desc_agg} - {""}  # pra pegar nomes normalizados
    todas_desc = list(desc_raw | desc_agg)
    try:
        _MAPA_NORMALIZACAO = _gerar_normalizacao(todas_desc, client, args.modelo)
    except Exception as e:
        print(f"   ⚠️  Normalização LLM falhou ({e}), usando cache existente")
        _MAPA_NORMALIZACAO, _ = _carregar_normalizacao()

    # Agrupamento de itens relacionados via LLM (cached)
    # Usa descrições pós-normalização + agregação (só itens)
    desc_para_agrupar = list({_normalizar(_agregar_descricao(d)) for d in desc_raw} - {""})
    try:
        _MAPA_AGRUPAMENTO = _gerar_agrupamento(desc_para_agrupar, client, args.modelo)
    except Exception as e:
        print(f"   ⚠️  Agrupamento LLM falhou ({e}), usando cache existente")
        _MAPA_AGRUPAMENTO, _ = _carregar_agrupamento()
    # Aba Evolução (pivot, sem gráficos — plots vão no HTML)
    construir_evolucao(wb, todos)

    # Aba Consolidado
    consolidado = []
    for mes, rows in todos.items():
        for r in rows:
            consolidado.append({"arquivo": mes, **r})

    if consolidado:
        ws = wb.create_sheet("Consolidado")
        ws.append(["Arquivo", "Descrição", "Valor", "Tipo"])
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="305496")
        for r in consolidado:
            ws.append([r["arquivo"], r.get("descricao", ""), r.get("valor"), r.get("tipo", "item")])
            ws.cell(row=ws.max_row, column=3).number_format = '#,##0.00;-#,##0.00'
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 65
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 12
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb.save(args.saida)
    print(f"\n✅ Excel salvo: {args.saida}")

    # HTML com gráficos interativos
    html_path = Path(args.saida).with_suffix(".html")
    img_map = {}
    for path in sorted(img_paths):
        mes_label = path.stem
        if mes_label in todos:
            img_map[mes_label] = path
    gerar_html(todos, img_map, html_path)


if __name__ == "__main__":
    main()
